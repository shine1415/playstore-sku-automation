"""
Excel file processing functionality for subscription management system.

This module provides functionality to parse subscription data from Excel files,
validate the data format, and generate template files for users.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import io
from decimal import Decimal, InvalidOperation

from data_models import (
    SubscriptionData, CountryPricing, ValidationResult, 
    validate_excel_row_data, validate_subscription_parameters,
    validate_country_pricing_data
)


class ExcelProcessor:
    """Handles Excel file processing for subscription data."""
    
    # Expected column names in Excel file
    REQUIRED_COLUMNS = [
        'Product ID', 'Package Name', 'Title', 'Description', 
        'Billing Period', 'Country Code', 'Currency', 'Price'
    ]
    
    OPTIONAL_COLUMNS = ['Trial Period', 'Grace Period']
    
    def __init__(self):
        """Initialize the Excel processor."""
        self.validation_errors = []
        self.validation_warnings = []
    
    def parse_excel_file(self, file_path: str) -> Tuple[List[SubscriptionData], ValidationResult]:
        """
        Parse subscription data from an Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (list of SubscriptionData objects, ValidationResult)
        """
        result = ValidationResult(is_valid=True)
        subscriptions = []
        
        try:
            # Read Excel file
            df = pd.read_excel(file_path, engine='openpyxl')
            
            # Validate file format
            format_validation = self._validate_file_format(df)
            if not format_validation.is_valid:
                result.errors.extend(format_validation.errors)
                result.warnings.extend(format_validation.warnings)
                result.is_valid = False
                return subscriptions, result
            
            # Group rows by Product ID to create subscription objects
            grouped_data = self._group_rows_by_product(df)
            
            # Process each product group
            for product_id, rows in grouped_data.items():
                subscription_result = self._create_subscription_from_rows(rows)
                
                if subscription_result[1].is_valid:
                    subscriptions.append(subscription_result[0])
                else:
                    result.errors.extend(subscription_result[1].errors)
                    result.warnings.extend(subscription_result[1].warnings)
                    result.is_valid = False
            
        except FileNotFoundError:
            result.add_error(f"Excel file not found: {file_path}")
        except Exception as e:
            result.add_error(f"Error reading Excel file: {str(e)}")
        
        return subscriptions, result 
    def _validate_file_format(self, df: pd.DataFrame) -> ValidationResult:
            """
            Validate the Excel file format and required columns.
            
            Args:
                df: Pandas DataFrame from Excel file
                
            Returns:
                ValidationResult with validation status
            """
            result = ValidationResult(is_valid=True)
            
            # Check if DataFrame is empty
            if df.empty:
                result.add_error("Excel file is empty")
                return result
            
            # Check for required columns
            missing_columns = []
            for col in self.REQUIRED_COLUMNS:
                if col not in df.columns:
                    missing_columns.append(col)
            
            if missing_columns:
                result.add_error(f"Missing required columns: {', '.join(missing_columns)}")
            
            # Check for unexpected columns (warn only)
            expected_columns = set(self.REQUIRED_COLUMNS + self.OPTIONAL_COLUMNS)
            unexpected_columns = set(df.columns) - expected_columns
            if unexpected_columns:
                result.add_warning(f"Unexpected columns found (will be ignored): {', '.join(unexpected_columns)}")
            
            # Check for completely empty rows
            empty_rows = df.isnull().all(axis=1).sum()
            if empty_rows > 0:
                result.add_warning(f"Found {empty_rows} completely empty rows (will be skipped)")
            
            return result
        
    def _group_rows_by_product(self, df: pd.DataFrame) -> Dict[str, List[Dict[str, Any]]]:
            """
            Group Excel rows by Product ID.
            
            Args:
                df: Pandas DataFrame from Excel file
                
            Returns:
                Dictionary mapping Product ID to list of row data
            """
            grouped_data = {}
            
            for index, row in df.iterrows():
                # Skip completely empty rows
                if row.isnull().all():
                    continue
                
                # Convert row to dictionary
                row_dict = row.to_dict()
                
                # Get product ID
                product_id = str(row_dict.get('Product ID', '')).strip()
                if not product_id:
                    continue
                
                if product_id not in grouped_data:
                    grouped_data[product_id] = []
                
                # Add row number for error reporting
                row_dict['_row_number'] = index + 2  # +2 because pandas is 0-indexed and Excel starts at row 1 (header)
                grouped_data[product_id].append(row_dict)
            
            return grouped_data 
    def _create_subscription_from_rows(self, rows: List[Dict[str, Any]]) -> Tuple[Optional[SubscriptionData], ValidationResult]:
            """
            Create a SubscriptionData object from grouped rows.
            
            Args:
                rows: List of row dictionaries for the same product
                
            Returns:
                Tuple of (SubscriptionData object or None, ValidationResult)
            """
            result = ValidationResult(is_valid=True)
            
            if not rows:
                result.add_error("No rows provided for subscription creation")
                return None, result
            
            # Validate each row
            for row in rows:
                row_validation = validate_excel_row_data(row, row.get('_row_number', 0))
                if not row_validation.is_valid:
                    result.errors.extend(row_validation.errors)
                    result.warnings.extend(row_validation.warnings)
                    result.is_valid = False
            
            if not result.is_valid:
                return None, result
            
            # Extract common subscription data from first row
            first_row = rows[0]
            
            try:
                # Create country pricing list from all rows
                country_pricing = []
                seen_countries = set()
                
                for row in rows:
                    country_code = str(row['Country Code']).strip().upper()
                    
                    # Check for duplicate countries
                    if country_code in seen_countries:
                        result.add_error(f"Duplicate country code '{country_code}' for product '{first_row['Product ID']}'")
                        result.is_valid = False
                        continue
                    
                    seen_countries.add(country_code)
                    
                    # Create CountryPricing object
                    try:
                        price_decimal = Decimal(str(row['Price']))
                        pricing = CountryPricing.from_decimal_price(
                            country_code=country_code,
                            currency_code=str(row['Currency']).strip().upper(),
                            price=float(price_decimal)
                        )
                        country_pricing.append(pricing)
                    except (ValueError, InvalidOperation) as e:
                        result.add_error(f"Invalid price data for {country_code}: {str(e)}")
                        result.is_valid = False
                
                if not result.is_valid:
                    return None, result
                
                # Create SubscriptionData object
                subscription = SubscriptionData(
                    product_id=str(first_row['Product ID']).strip(),
                    package_name=str(first_row.get('Package Name', '')).strip(),
                    title=str(first_row['Title']).strip(),
                    description=str(first_row['Description']).strip(),
                    billing_period=str(first_row['Billing Period']).strip(),
                    country_pricing=country_pricing,
                    trial_period=str(first_row.get('Trial Period', '')).strip() or None,
                    grace_period=str(first_row.get('Grace Period', '')).strip() or None
                )
                
                return subscription, result
                
            except Exception as e:
                result.add_error(f"Error creating subscription for product '{first_row.get('Product ID', 'unknown')}': {str(e)}")
                result.is_valid = False
                return None, result   
    def generate_template(self, output_path: str, include_sample_data: bool = True) -> bool:
            """
            Generate an Excel template file for subscription data input.
            
            Args:
                output_path: Path where the template file should be saved
                include_sample_data: Whether to include sample data rows
                
            Returns:
                True if template was created successfully, False otherwise
            """
            try:
                # Create workbook and worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = "Subscription Data"
                
                # Define headers
                headers = self.REQUIRED_COLUMNS + self.OPTIONAL_COLUMNS
                
                # Style for headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Add sample data if requested
                if include_sample_data:
                    sample_data = [
                        {
                            'Product ID': 'premium_monthly',
                            'Package Name': 'com.example.app',
                            'Title': 'Premium Monthly Subscription',
                            'Description': 'Access to premium features with monthly billing',
                            'Billing Period': 'P1M',
                            'Country Code': 'US',
                            'Currency': 'USD',
                            'Price': 9.99,
                            'Trial Period': 'P7D',
                            'Grace Period': 'P3D'
                        },
                        {
                            'Product ID': 'premium_monthly',
                            'Package Name': 'com.example.app',
                            'Title': 'Premium Monthly Subscription',
                            'Description': 'Access to premium features with monthly billing',
                            'Billing Period': 'P1M',
                            'Country Code': 'GB',
                            'Currency': 'GBP',
                            'Price': 7.99,
                            'Trial Period': 'P7D',
                            'Grace Period': 'P3D'
                        }
                    ]
                    
                    # Write sample data
                    for row_idx, data in enumerate(sample_data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            value = data.get(header, '')
                            ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save the workbook
                wb.save(output_path)
                return True
                
            except Exception as e:
                print(f"Error generating template: {str(e)}")
                return False
    
    def validate_excel_file(self, file_path: str) -> ValidationResult:
        """
        Validate an Excel file without parsing the full data.
        
        Args:
            file_path: Path to the Excel file to validate
            
        Returns:
            ValidationResult with validation status and any errors/warnings
        """
        result = ValidationResult(is_valid=True)
        
        try:
            # Check if file exists
            if not Path(file_path).exists():
                result.add_error(f"File does not exist: {file_path}")
                return result
            
            # Try to read the file
            df = pd.read_excel(file_path, engine='openpyxl')
            
            # Validate format
            format_result = self._validate_file_format(df)
            result.errors.extend(format_result.errors)
            result.warnings.extend(format_result.warnings)
            if not format_result.is_valid:
                result.is_valid = False
            
            # Validate each row
            for index, row in df.iterrows():
                if row.isnull().all():
                    continue
                
                row_dict = row.to_dict()
                row_dict['_row_number'] = index + 2
                
                row_result = validate_excel_row_data(row_dict, row_dict['_row_number'])
                result.errors.extend(row_result.errors)
                result.warnings.extend(row_result.warnings)
                if not row_result.is_valid:
                    result.is_valid = False
            
        except Exception as e:
            result.add_error(f"Error validating Excel file: {str(e)}")
        
        return result